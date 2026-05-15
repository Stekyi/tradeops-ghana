"""
Exports all analysis data to Excel workbooks in raw-data/ folder.
One workbook per report date, with one sheet per data category.
Run automatically by run_pipeline.py or manually: python scripts/export_to_excel.py
"""
import json
import logging
from datetime import date
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent))
from config import BASE_DIR, DATA_DIR, REPORTS_DIR

logger = logging.getLogger(__name__)

RAW_DATA_DIR = BASE_DIR / "raw-data"

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def _style_header(ws, header_row: int = 1):
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="CCCCCC")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = Border(bottom=Side(style="thin", color="EEEEEE"))
    ws.row_dimensions[header_row].height = 30


def export_report_to_excel(report_date: str = None) -> Path:
    if not HAS_PANDAS or not HAS_OPENPYXL:
        logger.error("pandas and openpyxl required. Run: pip install pandas openpyxl")
        return None

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    if report_date is None:
        dates = sorted([d.name for d in REPORTS_DIR.iterdir() if d.is_dir()], reverse=True)
        if not dates:
            logger.warning("No reports found. Run the pipeline first.")
            return None
        report_date = dates[0]

    analysis_path = REPORTS_DIR / report_date / "analysis.json"
    ghana_path = DATA_DIR / "ghana"
    comtrade_path = DATA_DIR / "comtrade"
    prices_path = DATA_DIR / "prices"

    if not analysis_path.exists():
        logger.error(f"No analysis.json found for {report_date}")
        return None

    with open(analysis_path, encoding="utf-8") as f:
        analysis = json.load(f)

    # Load raw data files
    ghana_files = sorted(ghana_path.glob("ghana_trade_*.json"), reverse=True)
    ghana_data = json.loads(ghana_files[0].read_text(encoding="utf-8")) if ghana_files else {}

    comtrade_files = sorted(comtrade_path.glob("comtrade_*.json"), reverse=True)
    comtrade_data = json.loads(comtrade_files[0].read_text(encoding="utf-8")) if comtrade_files else {}

    prices_files = sorted(prices_path.glob("prices_*.json"), reverse=True)
    prices_data = json.loads(prices_files[0].read_text(encoding="utf-8")) if prices_files else {}

    RAW_DATA_DIR.mkdir(parents=True, exist_ok=True)
    out_path = RAW_DATA_DIR / f"tradeops_{report_date}.xlsx"

    writer = pd.ExcelWriter(out_path, engine="openpyxl")

    # ── Sheet 1: Ghana Top Imports ─────────────────────────────────────
    top_imports = ghana_data.get("top_imports", [])
    if top_imports:
        df_imp = pd.DataFrame(top_imports)
        df_imp.columns = [c.replace("_", " ").title() for c in df_imp.columns]
        df_imp.insert(0, "Rank", range(1, len(df_imp) + 1))
        df_imp.to_excel(writer, sheet_name="Ghana Top Imports", index=False)
        logger.info("  [OK] Sheet: Ghana Top Imports")

    # ── Sheet 2: Ghana Non-Trad Exports ──────────────────────────────
    top_exports = ghana_data.get("top_exports_nontrad", [])
    if top_exports:
        df_exp = pd.DataFrame(top_exports)
        df_exp.columns = [c.replace("_", " ").title() for c in df_exp.columns]
        df_exp.insert(0, "Rank", range(1, len(df_exp) + 1))
        df_exp.to_excel(writer, sheet_name="Ghana Non-Trad Exports", index=False)
        logger.info("  [OK] Sheet: Ghana Non-Trad Exports")

    # ── Sheet 3: Ghana Macro Time-Series ─────────────────────────────
    macro_ts = ghana_data.get("macro_timeseries", {})
    if macro_ts:
        ts_rows = []
        all_years = sorted(set(yr for v in macro_ts.values() for yr in v.keys()), reverse=True)
        for yr in all_years:
            row = {"Year": yr}
            for indicator, values in macro_ts.items():
                row[indicator.replace("_", " ").title()] = values.get(yr)
            ts_rows.append(row)
        if ts_rows:
            pd.DataFrame(ts_rows).to_excel(writer, sheet_name="Macro Time-Series", index=False)
            logger.info("  [OK] Sheet: Macro Time-Series")

    # ── Sheet 4: Commodity Price Trends ──────────────────────────────
    prices = prices_data.get("prices", {})
    price_rows = []
    for comm, pdata in prices.items():
        for point in pdata.get("trend", []):
            price_rows.append({
                "Commodity": comm,
                "Year": point.get("year"),
                "Price (USD/MT)": point.get("price_usd_mt"),
                "Ghana Producer/Import Price (USD/MT)": pdata.get("ghana_producer_usd_mt") or pdata.get("import_price_usd_mt"),
                "Margin Potential (%)": pdata.get("margin_potential_pct"),
            })
    if price_rows:
        pd.DataFrame(price_rows).to_excel(writer, sheet_name="Commodity Prices", index=False)
        logger.info("  [OK] Sheet: Commodity Prices")

    # ── Sheet 5: Import Opportunities (Scored) ────────────────────────
    import_opps = analysis.get("import_opportunities", [])
    if import_opps:
        rows = []
        for o in import_opps:
            rows.append({
                "Commodity": o["commodity"],
                "Score (/100)": o["score"],
                "Ghana Annual Import ($M)": o["ghana_import_usd_m"],
                "YoY Growth (%)": o["yoy_growth_pct"],
                "Best Supplier": o["best_supplier"],
                "Best Price ($/MT)": o["best_price_usd_mt"],
                "Gross Margin (%)": o["margin_pct"],
                "Margin Flag": o["margin_flag"],
                "Budget Fit (GHC 100k)": "Yes" if o["budget_fit"] else "Tight",
                "Est. Order Qty (MT)": o["est_order_qty_mt"],
                "Est. Starting Order (GHC)": o["est_order_ghc"],
                "All Suppliers": "; ".join(f"{s['country']} ${s['price_usd_mt']}/MT" for s in o.get("all_suppliers", [])),
                "Score: Gap Size": o["score_breakdown"]["gap_size"],
                "Score: Growth": o["score_breakdown"]["growth"],
                "Score: Margin": o["score_breakdown"]["margin"],
                "Score: Logistics": o["score_breakdown"]["logistics"],
                "Score: Budget Fit": o["score_breakdown"]["budget_fit"],
            })
        pd.DataFrame(rows).sort_values("Score (/100)", ascending=False).to_excel(
            writer, sheet_name="Import Opportunities", index=False)
        logger.info("  [OK] Sheet: Import Opportunities")

    # ── Sheet 6: Export Opportunities (Scored) ────────────────────────
    export_opps = analysis.get("export_opportunities", [])
    if export_opps:
        rows = []
        for o in export_opps:
            rows.append({
                "Commodity": o["commodity"],
                "Score (/100)": o["score"],
                "Ghana Annual Export ($M)": o["ghana_export_usd_m"],
                "YoY Growth (%)": o["yoy_growth_pct"],
                "Best Target Market": o["best_market"],
                "Market Demand Growth (%)": o["market_demand_growth_pct"],
                "Price Premium": o["price_premium"],
                "Ghana Producer Price ($/MT)": o["ghana_producer_price_usd_mt"],
                "Gross Margin (%)": o["margin_pct"],
                "Margin Flag": o["margin_flag"],
                "Budget Fit (GHC 100k)": "Yes" if o["budget_fit"] else "Tight",
                "Est. Order Qty (MT)": o["est_order_qty_mt"],
                "All Target Markets": "; ".join(f"{m['country']} +{m['demand_growth_pct']}%" for m in o.get("all_markets", [])),
                "Score: Export Value": o["score_breakdown"]["export_value"],
                "Score: Growth": o["score_breakdown"]["growth"],
                "Score: Margin": o["score_breakdown"]["margin"],
                "Score: Market Demand": o["score_breakdown"]["market_demand"],
                "Score: Budget Fit": o["score_breakdown"]["budget_fit"],
            })
        pd.DataFrame(rows).sort_values("Score (/100)", ascending=False).to_excel(
            writer, sheet_name="Export Opportunities", index=False)
        logger.info("  [OK] Sheet: Export Opportunities")

    # ── Sheet 7: Top 5 Summary ────────────────────────────────────────
    top5 = analysis.get("top_5_overall", [])
    if top5:
        rows = []
        for i, o in enumerate(top5, 1):
            rows.append({
                "Rank": i,
                "Commodity": o["commodity"],
                "Direction": "Import" if o["type"] == "import_opportunity" else "Export",
                "Score": o["score"],
                "Gross Margin (%)": o["margin_pct"],
                "Starting Order (GHC)": o["est_order_ghc"],
                "Est. Order Qty (MT)": o["est_order_qty_mt"],
                "Best Partner": o.get("best_supplier") or o.get("best_market"),
            })
        pd.DataFrame(rows).to_excel(writer, sheet_name="Top 5 Opportunities", index=False)
        logger.info("  [OK] Sheet: Top 5 Opportunities")

    # ── Sheet 8: Raw Comtrade Commodity Flows ─────────────────────────
    comms = comtrade_data.get("commodities", {})
    flow_rows = []
    for comm, cdata in comms.items():
        for flow_key in ("exports", "imports"):
            for entry in cdata.get(flow_key, []):
                flow_rows.append({
                    "Commodity": comm,
                    "Flow": flow_key.title(),
                    "Partner Country": entry.get("partner"),
                    "Value (USD)": entry.get("value_usd"),
                    "Quantity (MT)": entry.get("qty_mt"),
                    "Year": entry.get("year"),
                    "Unit Price ($/MT)": round(entry["value_usd"] / entry["qty_mt"], 2)
                        if entry.get("qty_mt") and entry.get("value_usd") else None,
                })
    if flow_rows:
        pd.DataFrame(flow_rows).to_excel(writer, sheet_name="Commodity Trade Flows", index=False)
        logger.info("  [OK] Sheet: Commodity Trade Flows")

    writer.close()

    # Apply styling
    wb = load_workbook(out_path)
    for ws in wb.worksheets:
        _style_header(ws)
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
        ws.freeze_panes = "A2"
    wb.save(out_path)

    logger.info(f"Excel workbook saved -> {out_path}")
    return out_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    path = export_report_to_excel()
    if path:
        print(f"Saved: {path}")
